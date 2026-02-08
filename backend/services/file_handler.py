"""
Temporary file storage and content extraction for chat attachments.

Handles:
- Temp storage of uploaded files keyed by upload_id
- Content extraction / base64 encoding for building Claude API content blocks
- Automatic cleanup after retrieval

Supported file types:
- Images (jpeg, png, gif, webp) → sent as native Claude image blocks
- PDFs → sent as native Claude document blocks
- XLSX → parsed to CSV text via openpyxl
- CSV / plain text → inlined as text blocks
"""

from __future__ import annotations

import base64
import csv
import io
import logging
import mimetypes
import threading
import time
import uuid
from dataclasses import dataclass, field
from typing import Any

logger = logging.getLogger(__name__)

# Max file size: 10 MB
MAX_FILE_SIZE: int = 10 * 1024 * 1024

# TTL for stored files: 30 minutes
FILE_TTL_SECONDS: int = 30 * 60

# MIME types Claude accepts natively as image blocks
NATIVE_IMAGE_MIMES: frozenset[str] = frozenset({
    "image/jpeg",
    "image/png",
    "image/gif",
    "image/webp",
})

# MIME type for native PDF support
PDF_MIME: str = "application/pdf"

# Text-like MIME prefixes / exact matches we inline as text
TEXT_MIMES: frozenset[str] = frozenset({
    "text/plain",
    "text/csv",
    "text/markdown",
    "text/html",
    "application/json",
    "application/xml",
    "text/xml",
})

XLSX_MIME: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLS_MIME: str = "application/vnd.ms-excel"


@dataclass
class StoredFile:
    """A temporarily stored uploaded file."""
    upload_id: str
    filename: str
    mime_type: str
    size: int
    data: bytes
    created_at: float = field(default_factory=time.time)


# In-memory store — keyed by upload_id
_store: dict[str, StoredFile] = {}
_lock: threading.Lock = threading.Lock()


def store_file(filename: str, data: bytes, content_type: str | None = None) -> StoredFile:
    """
    Store an uploaded file and return its metadata.

    Args:
        filename: Original filename from the upload
        data: Raw file bytes
        content_type: MIME type (guessed from filename if not provided)

    Returns:
        StoredFile with generated upload_id

    Raises:
        ValueError: If file exceeds MAX_FILE_SIZE
    """
    if len(data) > MAX_FILE_SIZE:
        raise ValueError(f"File exceeds maximum size of {MAX_FILE_SIZE // (1024 * 1024)} MB")

    mime: str = content_type or mimetypes.guess_type(filename)[0] or "application/octet-stream"
    upload_id: str = str(uuid.uuid4())

    stored = StoredFile(
        upload_id=upload_id,
        filename=filename,
        mime_type=mime,
        size=len(data),
        data=data,
    )

    with _lock:
        _store[upload_id] = stored
        _cleanup_expired()

    logger.info("Stored file %s (%s, %d bytes) as %s", filename, mime, len(data), upload_id)
    return stored


def retrieve_file(upload_id: str) -> StoredFile | None:
    """Retrieve a stored file by upload_id (non-destructive)."""
    with _lock:
        return _store.get(upload_id)


def remove_file(upload_id: str) -> None:
    """Remove a stored file after it's been consumed."""
    with _lock:
        _store.pop(upload_id, None)


def _cleanup_expired() -> None:
    """Remove files older than FILE_TTL_SECONDS. Must be called under _lock."""
    now: float = time.time()
    expired: list[str] = [
        uid for uid, sf in _store.items()
        if now - sf.created_at > FILE_TTL_SECONDS
    ]
    for uid in expired:
        del _store[uid]
    if expired:
        logger.info("Cleaned up %d expired upload(s)", len(expired))


# ---------------------------------------------------------------------------
# Content extraction — builds Claude API content blocks
# ---------------------------------------------------------------------------

def build_claude_content_blocks(
    stored_files: list[StoredFile],
) -> list[dict[str, Any]]:
    """
    Convert a list of StoredFiles into Claude API content blocks.

    - Images → {"type": "image", "source": {"type": "base64", ...}}
    - PDFs   → {"type": "document", "source": {"type": "base64", ...}}
    - XLSX   → {"type": "text", "text": "Contents of <filename>:\n<csv>"}
    - CSV/text → {"type": "text", "text": "Contents of <filename>:\n<text>"}
    """
    blocks: list[dict[str, Any]] = []

    for sf in stored_files:
        mime: str = sf.mime_type

        if mime in NATIVE_IMAGE_MIMES:
            blocks.append(_image_block(sf))
        elif mime == PDF_MIME:
            blocks.append(_pdf_block(sf))
        elif mime in (XLSX_MIME, XLS_MIME):
            blocks.append(_xlsx_to_text_block(sf))
        elif _is_text_mime(mime) or sf.filename.endswith(".csv"):
            blocks.append(_text_file_block(sf))
        else:
            # Best-effort: try to decode as text, fall back to skipping
            blocks.append(_best_effort_text_block(sf))

    return blocks


def _image_block(sf: StoredFile) -> dict[str, Any]:
    b64: str = base64.standard_b64encode(sf.data).decode("ascii")
    return {
        "type": "image",
        "source": {
            "type": "base64",
            "media_type": sf.mime_type,
            "data": b64,
        },
    }


def _pdf_block(sf: StoredFile) -> dict[str, Any]:
    b64: str = base64.standard_b64encode(sf.data).decode("ascii")
    return {
        "type": "document",
        "source": {
            "type": "base64",
            "media_type": "application/pdf",
            "data": b64,
        },
    }


def _xlsx_to_text_block(sf: StoredFile) -> dict[str, Any]:
    """Parse XLSX to CSV text using openpyxl."""
    try:
        import openpyxl

        wb: openpyxl.Workbook = openpyxl.load_workbook(
            io.BytesIO(sf.data), read_only=True, data_only=True,
        )
        parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            output = io.StringIO()
            writer = csv.writer(output)

            for row in ws.iter_rows(values_only=True):
                writer.writerow([
                    str(cell) if cell is not None else ""
                    for cell in row
                ])

            sheet_csv: str = output.getvalue()
            if len(wb.sheetnames) > 1:
                parts.append(f"--- Sheet: {sheet_name} ---\n{sheet_csv}")
            else:
                parts.append(sheet_csv)

        wb.close()
        text: str = "\n".join(parts)

        # Truncate very large spreadsheets to ~200k chars to stay within token limits
        max_chars: int = 200_000
        if len(text) > max_chars:
            text = text[:max_chars] + f"\n\n[Truncated — showing first {max_chars:,} characters of {len(text):,} total]"

        return {
            "type": "text",
            "text": f"Contents of {sf.filename}:\n\n{text}",
        }

    except Exception as e:
        logger.warning("Failed to parse XLSX %s: %s", sf.filename, e)
        return {
            "type": "text",
            "text": f"[Attached file: {sf.filename} — could not parse Excel file: {e}]",
        }


def _text_file_block(sf: StoredFile) -> dict[str, Any]:
    """Decode file as UTF-8 text."""
    try:
        text: str = sf.data.decode("utf-8")
    except UnicodeDecodeError:
        text = sf.data.decode("latin-1")

    max_chars: int = 200_000
    if len(text) > max_chars:
        text = text[:max_chars] + f"\n\n[Truncated — showing first {max_chars:,} characters of {len(text):,} total]"

    return {
        "type": "text",
        "text": f"Contents of {sf.filename}:\n\n{text}",
    }


def _best_effort_text_block(sf: StoredFile) -> dict[str, Any]:
    """Try to decode as text; if it fails, return a placeholder."""
    try:
        text: str = sf.data.decode("utf-8")
        max_chars: int = 200_000
        if len(text) > max_chars:
            text = text[:max_chars] + f"\n\n[Truncated]"
        return {
            "type": "text",
            "text": f"Contents of {sf.filename}:\n\n{text}",
        }
    except UnicodeDecodeError:
        return {
            "type": "text",
            "text": f"[Attached file: {sf.filename} ({sf.mime_type}, {sf.size:,} bytes) — binary format not supported for inline display]",
        }


def _is_text_mime(mime: str) -> bool:
    """Check if a MIME type is text-like."""
    return mime in TEXT_MIMES or mime.startswith("text/")
